from typing import Any, Generator

from openpyxl import load_workbook

from app.services import InvalidAmountError


def get_articles_and_prices(
    excel_path: str
) -> Generator[tuple[int, int], Any, Any]:

    wb = load_workbook(excel_path)
    if not (sheet := wb.active):
        return
    
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row[0] or not row[1]:
            continue
        yield int(row[0]), int(row[1])


def get_skus_and_amount(
    excel_path: str
) -> Generator[tuple[str, int], Any, Any]:
    wb = load_workbook(excel_path)
    if not (sheet := wb.active):
        return

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue

        sku = str(row[0]).rstrip('.0')
        raw_amount = row[1]

        try:
            amount = int(raw_amount)
            if amount < 0:
                raise ValueError("Отрицательное количество")
        except (ValueError, TypeError):
            raise InvalidAmountError(sku=sku, amount=raw_amount)

        yield sku, amount
