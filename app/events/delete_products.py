import requests
import PySimpleGUI as sg
from openpyxl import load_workbook

from app.events.utils import get_shop_warehouses, get_product_card
from app.interface.colors import WHITE_COLOR, BLACK_COLOR, BLUE_COLOR


def remove_product_rest(
    api_key: str,
    excel_path: str,
) -> bool:
    TRASH_URL = (
        'https://content-api.wildberries.ru/content/v2/cards/delete/trash'
    )

    is_products = False
    errors = False
    error_message = None

    headers = {"Authorization": api_key}

    warehouses_data = get_shop_warehouses(api_key)
    warehouses_id = [*map(lambda x: x['id'], warehouses_data)]

    wb = load_workbook(excel_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, max_col=0, values_only=True):

        article = row[0]
        product_data = get_product_card(api_key, article)
        if not product_data['cards']:
            continue
        is_products = True
        skuses = [*map(lambda x: x['skus'], product_data['cards'][0]['sizes'])]
        nmID = product_data["cards"][0]["nmID"]

        for skus in skuses:

            params = {
                'skus': skus,
            }

            for id in warehouses_id:

                STOCKS_URL = (
                    'https://marketplace-api.wildberries.ru/'
                    f'api/v3/stocks/{id}'
                )

                result = requests.delete(
                    STOCKS_URL,
                    headers=headers,
                    json=params,
                )

                if result.status_code != 204:
                    errors = True
                    error_message = 'Некоторые товары не были удалены.'

        result = requests.post(
            TRASH_URL,
            headers=headers,
            json={'nmIDs': [nmID]}
        )

        if result.status_code != 200:
            errors = True
            error_message = 'Некоторые товары не удалось переместить в корзину'

    if not is_products:
        errors = True
        error_message = 'В текущих артикулах не найдены товары для удаления'

    return errors, error_message


def process_product_delete(event, file_path, shops: dict, window):
    shop_name = event.removeprefix('DELETE_POPUP')

    api_key = shops[shop_name]

    path = file_path[shop_name]

    del file_path[shop_name]

    errors, error_message = remove_product_rest(api_key, path)

    if errors:
        sg.popup(
            error_message,
            title='Результаты запросов',
            background_color=WHITE_COLOR,
            text_color=BLACK_COLOR,
            button_color=BLUE_COLOR
        )

    else:
        sg.popup(
            'Все прошло успешно',
            title='Результаты запросов',
            background_color=WHITE_COLOR,
            text_color=BLACK_COLOR,
            button_color=BLUE_COLOR
        )
