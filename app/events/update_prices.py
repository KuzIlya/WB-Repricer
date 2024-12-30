import requests
import logging
import PySimpleGUI as sg
from openpyxl import load_workbook

from app.interface.colors import WHITE_COLOR, BLACK_COLOR, BLUE_COLOR
from app.events.utils import get_product_card

logging.basicConfig(
    filename='price_update.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


def update_prices(api_key, excel_path):

    URL = 'https://discounts-prices-api.wildberries.ru/api/v2/upload/task'

    wb = load_workbook(excel_path)
    sheet = wb.active

    success_count = 0
    failure_count = 0

    goods_data = []
    rows_to_delete = []

    for row in sheet.iter_rows(min_row=2, max_col=0, values_only=True):
        article, price = row
        if article and price:
            product_data = get_product_card(api_key, article)
            if (
                product_data
                and "cards" in product_data
                and product_data["cards"]
            ):
                nmID = product_data["cards"][0]["nmID"]
                goods_data.append({"nmID": nmID, "price": price})
                rows_to_delete.append(article)
            else:
                logging.warning(f"Не удалось найти товар с артикулом {article}.")

    if not goods_data:
        logging.info("Нет данных для обновления.")
        print("Нет данных для обновления.")
        return

    batch_size = 1000
    batches = [
        goods_data[i:i + batch_size]
        for i in range(0, len(goods_data), batch_size)
    ]

    headers = {"Authorization": api_key}

    try:
        for batch in batches:
            payload = {"data": batch}

            response = requests.post(URL, headers=headers, json=payload)
            response_data = response.json()

            if (
                response.status_code == 200
                and not response_data.get("error", True)
            ):
                success_count += len(batch)
                logging.info(f"Успешно обновлено {len(batch)} товаров.")
                for item in batch:
                    for row in sheet.iter_rows(min_row=2, values_only=False):
                        if row[0].value == item["nmID"]:
                            sheet.delete_rows(row[0].row, 1)
                            break
            else:
                failure_count += len(batch)
                error_text = response_data.get('errorText', 'Неизвестная ошибка')
                logging.error(
                    f"Ошибка при обновлении: {error_text} для {len(batch)} товаров."
                )
    except Exception as e:
        logging.exception(f"Произошла ошибка при выполнении запроса: {e}")
        print(f"Произошла ошибка при выполнении запроса: {e}")
        failure_count += len(goods_data)

    wb.save(excel_path)

    return (success_count, failure_count)

def process_products_prices(event, file_path, shops: dict, window) -> None:
    shop_name = event.removeprefix('REFRESH')

    api_key = shops[shop_name]

    path = file_path[shop_name]

    del file_path[shop_name]

    result = update_prices(api_key, path)

    if not result:
        sg.popup(
            'Нет данных для обновления',
            title='Результаты запросов',
            background_color=WHITE_COLOR,
            text_color=BLACK_COLOR,
            button_color=BLUE_COLOR
        )

    else:

        sg.popup(
            (f'Измененных товаров: {result[0]}\n'
             f'Неизмененных товаров: {result[1]}'),
            title='Результаты запросов',
            background_color=WHITE_COLOR,
            text_color=BLACK_COLOR,
            button_color=BLUE_COLOR
        )

    refresh_button = window.find_element('REFRESH' + shop_name)
    delete_popup_button = window.find_element('DELETE_POPUP' + shop_name)

    refresh_button.update(disabled=True)
    delete_popup_button.update(disabled=True)
