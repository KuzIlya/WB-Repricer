import requests
import time
from openpyxl import load_workbook
from typing import Optional
import PySimpleGUI as sg

from app.interface.colors import WHITE_COLOR, BLACK_COLOR, BLUE_COLOR


def add_file(event, values, file_path, window: sg.Window) -> None:
    shop_name = event.removeprefix('ADD_FILE')
    file_path[shop_name] = values[event]

    refresh_button = window.find_element('REFRESH' + shop_name)

    refresh_button.update(disabled=False)

    window.refresh()


def get_product_card(
    api_key: str,
    article: str,
    retries: int = 3
) -> Optional[dict]:
    """
    Получает данные карточки товара по артикулу.

    :param api_key: API-ключ для авторизации.
    :param article: Артикул товара.
    :param retries: Количество попыток при ошибках.
    :return: JSON с данными карточки товара или None в случае неудачи.
    """
    url = "https://content-api.wildberries.ru/content/v2/get/cards/list"
    headers = {
        'Authorization': api_key,
        'Content-Type': 'application/json'
    }
    params = {
        "settings": {
            "cursor": {
                "limit": 100
            },
            "filter": {
                "withPhoto": 1,
                "textSearch": str(article)
            }
        }
    }

    for attempt in range(retries):
        try:
            response = requests.post(
                url,
                headers=headers,
                json=params,
                timeout=60
            )
            if response.status_code == 200:
                return response.json()
            else:
                print(f"Ошибка: {response.status_code}, {response.text}")
        except requests.exceptions.RequestException as e:
            print(f"Попытка {attempt + 1}/{retries} завершилась ошибкой: {e}")
            if attempt < retries - 1:
                time.sleep(2)
            else:
                raise e
    return None


def update_prices(api_key, excel_path):

    URL = 'https://discounts-prices-api.wildberries.ru/api/v2/upload/task'

    wb = load_workbook(excel_path)
    sheet = wb.active

    success_count = 0
    failure_count = 0

    goods_data = []
    rows_to_delete = []

    for row in sheet.iter_rows(min_row=2, max_col=0, values_only=True):
        article, price = row
        if article and price:
            product_data = get_product_card(api_key, article)
            if (
                product_data
                and "cards" in product_data
                and product_data["cards"]
            ):
                nmID = product_data["cards"][0]["nmID"]
                goods_data.append({"nmID": nmID, "price": price})
                rows_to_delete.append(article)
            else:
                print(f"Не удалось найти товар с артикулом {article}.")

    if not goods_data:
        print("Нет данных для обновления.")
        return

    batch_size = 1000
    batches = [
        goods_data[i:i + batch_size]
        for i in range(0, len(goods_data), batch_size)
    ]

    headers = {"Authorization": api_key}

    try:
        for batch in batches:
            payload = {"data": batch}

            response = requests.post(URL, headers=headers, json=payload)
            response_data = response.json()

            if (
                response.status_code == 200
                and not response_data.get("error", True)
            ):
                success_count += len(batch)
                for item in batch:
                    for row in sheet.iter_rows(min_row=2, values_only=False):
                        if row[0].value == item["nmID"]:
                            sheet.delete_rows(row[0].row, 1)
                            break
            else:
                failure_count += len(batch)
                print(
                    'Ошибка: '
                    f"{response_data.get('errorText', 'Неизвестная ошибка')}"
                )
    except Exception as e:
        print(f"Произошла ошибка при выполнении запроса: {e}")
        failure_count += len(goods_data)

    wb.save(excel_path)

    return (success_count, failure_count)


def process_file(event, file_path, shops: dict, window) -> None:
    shop_name = event.removeprefix('REFRESH')

    api_key = shops[shop_name]

    path = file_path[shop_name]

    del file_path[shop_name]

    result = update_prices(api_key, path)

    if not result:
        sg.popup(
            'Нет данных для обновления',
            title='Результаты запросов',
            background_color=WHITE_COLOR,
            text_color=BLACK_COLOR,
            button_color=BLUE_COLOR
        )

    else:

        sg.popup(
            (f'Измененных товаров: {result[0]}\n'
             f'Неизмененных товаров: {result[1]}'),
            title='Результаты запросов',
            background_color=WHITE_COLOR,
            text_color=BLACK_COLOR,
            button_color=BLUE_COLOR
        )

    refresh_button = window.find_element('REFRESH' + shop_name)
    refresh_button.update(disabled=True)
